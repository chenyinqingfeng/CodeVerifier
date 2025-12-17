"""
Excel导出器模块（重构版）
支持扫码日志和生产报告的Excel导出，包含统计汇总和明细数据
"""

from typing import Optional, List, Dict, Any
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .export_query_service import ExportQueryService


class ExcelExporter:
    """Excel导出器（重构版）"""

    def __init__(self, db_manager):
        """
        初始化导出器

        Args:
            db_manager: DatabaseManager实例
        """
        self.db = db_manager
        self.query_service = ExportQueryService(db_manager)

        # Excel样式定义
        self.style_header = {
            'font': Font(name='微软雅黑', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': self._create_border()
        }

        self.style_data = {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': self._create_border()
        }

        self.style_success = {
            'font': Font(name='微软雅黑', size=10),
            'fill': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': self._create_border()
        }

        self.style_fail = {
            'font': Font(name='微软雅黑', size=10),
            'fill': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': self._create_border()
        }

        self.style_gray = {
            'font': Font(name='微软雅黑', size=10, color='666666'),
            'fill': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': self._create_border()
        }

        # 为日志导出方法提供的单独样式属性
        self.header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.border = self._create_border()

    def _create_border(self):
        """创建单元格边框"""
        side = Side(style='thin', color='D0D0D0')
        return Border(left=side, right=side, top=side, bottom=side)

    def _apply_style(self, cell, style: Dict[str, Any]):
        """应用样式到单元格"""
        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'alignment' in style:
            cell.alignment = style['alignment']
        if 'border' in style:
            cell.border = style['border']

    def _auto_adjust_column_width(self, sheet):
        """自动调整列宽（超宽松版，舒适阅读）"""
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        # 中文字符计数为2.2个宽度（更宽）
                        cell_str = str(cell.value)
                        chinese_count = sum(1 for c in cell_str if '\u4e00' <= c <= '\u9fff')
                        english_count = len(cell_str) - chinese_count
                        char_length = chinese_count * 2.2 + english_count * 1.2
                        max_length = max(max_length, char_length)
                except Exception:
                    pass

            # 超宽松设置：最小宽度15，额外空间+6，最大宽度100
            adjusted_width = max(15, min(max_length + 6, 100))
            sheet.column_dimensions[column].width = adjusted_width

    # ==================== 扫码日志导出 ====================

    def export_scan_logs(
        self,
        output_path: str,
        customer_ids: Optional[List[int]] = None,
        container_ids: Optional[List[str]] = None,
        batch_ids: Optional[List[int]] = None,
        start_time: Optional[str] = None,
        end_time: Optional[str] = None,
        scan_result: Optional[str] = None
    ) -> bool:
        """
        导出扫码日志到Excel

        Args:
            output_path: 输出文件路径
            customer_ids: 客户ID列表
            container_ids: 货柜批号列表
            batch_ids: 批次ID列表
            start_time: 开始时间 (YYYY-MM-DD HH:MM:SS)
            end_time: 结束时间 (YYYY-MM-DD HH:MM:SS)
            scan_result: 扫描结果筛选 ('pass'/'fail'/None)

        Returns:
            是否导出成功
        """
        try:
            print(f"[INFO] 开始导出扫码日志...")

            # 查询数据
            logs = self.query_service.query_scan_logs(
                customer_ids=customer_ids,
                container_ids=container_ids,
                batch_ids=batch_ids,
                start_time=start_time,
                end_time=end_time,
                scan_result=scan_result
            )

            if not logs:
                print("[WARN] 没有符合条件的扫码日志数据")
                return False

            # 查询统计信息
            stats = self.query_service.get_scan_logs_statistics(
                customer_ids=customer_ids,
                container_ids=container_ids,
                batch_ids=batch_ids,
                start_time=start_time,
                end_time=end_time
            )

            # 创建工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认Sheet

            # 创建统计汇总Sheet
            self._create_scan_log_summary_sheet(wb, stats)

            # 创建日志详细Sheet
            self._create_scan_log_detail_sheet(wb, logs)

            # 保存文件
            wb.save(output_path)
            print(f"[OK] 扫码日志导出成功: {output_path}")
            print(f"     共导出 {len(logs)} 条记录")

            return True

        except Exception as e:
            print(f"[ERROR] 导出扫码日志失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _create_scan_log_summary_sheet(self, wb, stats: Dict[str, Any]):
        """创建扫码日志统计汇总Sheet"""
        ws = wb.create_sheet("统计汇总", 0)

        # 标题行
        ws['A1'] = '扫码日志统计汇总'
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.font = Font(name='微软雅黑', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 汇总数据
        row = 3
        summary_data = [
            ('总扫码次数', stats['total_count']),
            ('成功次数', stats['success_count']),
            ('失败次数', stats['fail_count']),
            ('成功率', f"{stats['success_rate']}%")
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            self._apply_style(ws[f'A{row}'], {
                'font': Font(name='微软雅黑', size=10, bold=True),
                'alignment': Alignment(horizontal='right'),
                'border': self._create_border()
            })
            self._apply_style(ws[f'B{row}'], self.style_data)
            row += 1

        # 失败原因分类
        if stats['error_breakdown']:
            row += 1
            ws[f'A{row}'] = '失败原因分类统计'
            ws.merge_cells(f'A{row}:B{row}')
            header_cell = ws[f'A{row}']
            self._apply_style(header_cell, self.style_header)
            row += 1

            # 表头
            ws[f'A{row}'] = '错误原因'
            ws[f'B{row}'] = '次数'
            self._apply_style(ws[f'A{row}'], self.style_header)
            self._apply_style(ws[f'B{row}'], self.style_header)
            row += 1

            # 错误数据
            for error_msg, count in stats['error_breakdown'].items():
                ws[f'A{row}'] = error_msg
                ws[f'B{row}'] = count
                self._apply_style(ws[f'A{row}'], self.style_data)
                self._apply_style(ws[f'B{row}'], self.style_data)
                row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_scan_log_detail_sheet(self, wb, logs: List[Dict[str, Any]]):
        """创建扫码日志详细Sheet（适配新数据结构）"""
        ws = wb.create_sheet("扫码日志详细")

        # 表头 - 适配新的scan_logs表结构
        headers = ['扫码时间', '扫码枪', '扫码数据', '扫码结果', '结果说明', '操作员']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.style_header)

        # 数据行
        for row_idx, log in enumerate(logs, start=2):
            ws.cell(row=row_idx, column=1, value=log['scan_time'])
            ws.cell(row=row_idx, column=2, value=log['scanner_port'])
            ws.cell(row=row_idx, column=3, value=log['scan_data'])
            ws.cell(row=row_idx, column=4, value=log['scan_result'])
            ws.cell(row=row_idx, column=5, value=log['result_message'])
            ws.cell(row=row_idx, column=6, value=log['operator_name'])

            # 根据结果应用样式
            scan_result = log['scan_result']
            if scan_result == '验证通过':
                style = self.style_success
            elif scan_result in ['验证失败', '不在批次', '二码不一致']:
                style = self.style_fail
            elif scan_result == '重复扫码':
                style = self.style_gray
            else:
                style = self.style_data

            for col_idx in range(1, 7):
                self._apply_style(ws.cell(row=row_idx, column=col_idx), style)

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 添加筛选器
        ws.auto_filter.ref = ws.dimensions

        # 自动调整列宽
        self._auto_adjust_column_width(ws)

    # ==================== 生产报告导出 ====================

    def export_production_report(
        self,
        output_path: str,
        customer_ids: Optional[List[int]] = None,
        container_ids: Optional[List[str]] = None,
        batch_ids: Optional[List[int]] = None,
        match_status: Optional[str] = None
    ) -> bool:
        """
        导出生产报告到Excel

        Args:
            output_path: 输出文件路径
            customer_ids: 客户ID列表
            container_ids: 货柜批号列表
            batch_ids: 批次ID列表
            match_status: 匹配状态 ('matched'/'unmatched'/None)

        Returns:
            是否导出成功
        """
        try:
            print(f"[INFO] 开始导出生产报告...")

            # 查询数据
            barcodes = self.query_service.query_production_report(
                customer_ids=customer_ids,
                container_ids=container_ids,
                batch_ids=batch_ids,
                match_status=match_status
            )

            if not barcodes:
                print("[WARN] 没有符合条件的生产报告数据")
                return False

            # 查询统计信息
            stats = self.query_service.get_production_report_statistics(
                customer_ids=customer_ids,
                container_ids=container_ids,
                batch_ids=batch_ids
            )

            # 创建工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # 创建生产报告概览Sheet
            self._create_production_summary_sheet(wb, stats)

            # 创建条码明细Sheet
            self._create_barcode_detail_sheet(wb, barcodes)

            # 保存文件
            wb.save(output_path)
            print(f"[OK] 生产报告导出成功: {output_path}")
            print(f"     共导出 {len(barcodes)} 条条码记录")

            return True

        except Exception as e:
            print(f"[ERROR] 导出生产报告失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _create_production_summary_sheet(self, wb, stats: Dict[str, Any]):
        """创建生产报告概览Sheet"""
        ws = wb.create_sheet("生产报告概览", 0)

        # 标题行
        ws['A1'] = '生产报告概览'
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.font = Font(name='微软雅黑', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 汇总数据
        row = 3
        summary_data = [
            ('客户数量', stats['customer_count']),
            ('货柜数量', stats['container_count']),
            ('批次数量', stats['batch_count']),
            ('', ''),
            ('总条码数', stats['total_barcode_count']),
            ('已扫数量', stats['scanned_count']),
            ('未扫数量', stats['unscanned_count']),
            ('扫码率', f"{stats['scan_rate']}%"),
            ('', ''),
            ('已打印数量', stats['printed_count']),
            ('未打印数量', stats['unprinted_count'])
        ]

        for label, value in summary_data:
            if label:  # 跳过空行
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                self._apply_style(ws[f'A{row}'], {
                    'font': Font(name='微软雅黑', size=10, bold=True),
                    'alignment': Alignment(horizontal='right'),
                    'border': self._create_border()
                })
                self._apply_style(ws[f'B{row}'], self.style_data)
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_barcode_detail_sheet(self, wb, barcodes: List[Dict[str, Any]]):
        """创建条码明细Sheet"""
        ws = wb.create_sheet("条码明细")

        # 表头
        headers = ['客户名称', '货柜批号', '批次名称', '生产条码', '扫码状态', '扫码时间', '重复扫码次数', '打印状态', '打印时间']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.style_header)

        # 数据行
        for row_idx, barcode in enumerate(barcodes, start=2):
            ws.cell(row=row_idx, column=1, value=barcode['customer_name'])
            ws.cell(row=row_idx, column=2, value=barcode['container_id'])
            ws.cell(row=row_idx, column=3, value=barcode['batch_name'])
            ws.cell(row=row_idx, column=4, value=barcode['barcode'])
            ws.cell(row=row_idx, column=5, value=barcode['is_matched'])
            ws.cell(row=row_idx, column=6, value=barcode['scan_time'])
            ws.cell(row=row_idx, column=7, value=barcode.get('scan_count', '-'))
            ws.cell(row=row_idx, column=8, value=barcode['is_printed'])
            ws.cell(row=row_idx, column=9, value=barcode['print_time'])

            # 根据扫码状态应用样式
            style = self.style_success if barcode['is_matched'] == '已扫' else self.style_gray
            for col_idx in range(1, 10):
                self._apply_style(ws.cell(row=row_idx, column=col_idx), style)

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 添加筛选器
        ws.auto_filter.ref = ws.dimensions

        # 自动调整列宽
        self._auto_adjust_column_width(ws)

    # ==================== 文件名生成工具 ====================

    @staticmethod
    def generate_scan_log_filename(start_time: str, end_time: str) -> str:
        """
        生成扫码日志文件名

        Args:
            start_time: 开始时间 (YYYY-MM-DD HH:MM:SS)
            end_time: 结束时间 (YYYY-MM-DD HH:MM:SS)

        Returns:
            文件名（不含路径）：扫码日志_起始YYYYMMDD_HHMM_截至YYYYMMDD_HHMM.xlsx
        """
        try:
            start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
            end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")

            start_str = start_dt.strftime("%Y%m%d_%H%M")
            end_str = end_dt.strftime("%Y%m%d_%H%M")

            return f"扫码日志_{start_str}_{end_str}.xlsx"
        except Exception:
            # 如果解析失败，使用当前时间
            now = datetime.now().strftime("%Y%m%d_%H%M")
            return f"扫码日志_{now}.xlsx"

    @staticmethod
    def generate_production_report_filename(
        customer_names: List[str],
        container_ids: List[str]
    ) -> str:
        """
        生成生产报告文件名

        Args:
            customer_names: 客户名称列表
            container_ids: 货柜批号列表

        Returns:
            文件名：生产报告_客户1&客户2_货柜1&货柜2_YYYYMMDD_HHMM.xlsx
        """
        # 客户名称处理（最多显示3个）
        if not customer_names:
            customer_str = "全部客户"
        elif len(customer_names) <= 3:
            customer_str = "&".join(customer_names)
        else:
            customer_str = "&".join(customer_names[:3]) + f"等{len(customer_names)}个"

        # 货柜批号处理（最多显示3个）
        if not container_ids:
            container_str = "全部货柜"
        elif len(container_ids) <= 3:
            container_str = "&".join(container_ids)
        else:
            container_str = "&".join(container_ids[:3]) + f"等{len(container_ids)}个"

        # 当前时间
        now = datetime.now().strftime("%Y%m%d_%H%M")

        # 过滤Windows不支持的文件名特殊字符
        filename = f"生产报告_{customer_str}_{container_str}_{now}.xlsx"
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in invalid_chars:
            filename = filename.replace(char, '_')

        return filename
